import os
import sys
import pdftotext
from openpyxl import load_workbook
class fileList:
    def __init__(self, PDFs, Excel):
        self.PDFs = PDFs
        self.Excel = Excel

#### Function Definitions ####

# This Function check the current folder for the files  needed to run this appication (PDF and Excel)
# Then Check If Files Are There
def checkForFiles():

    #This Command Is Needed For The App, Gets The Current Directory Of The App
    #os.chdir(os.path.dirname(sys.argv[0]))

    PDFs = [f for f in os.listdir('.') if f.endswith('.pdf')]
    Excel = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    if not Excel:
        print("Template Not Found!")
        sys.exit(1)
    if not PDFs:
        print("PDFs Not Found!")
        sys.exit(1)

    #Gives The User A Choice When There Is More Than One .xlsx Detected
    if len(Excel) > 1:
        print("Please Clarify Which One Is The Template")
        for i, name in enumerate(Excel):
            print i, ": ", name
        while True:
            try:
                ExcelChoice = int(raw_input("Which Number?: "))
            except ValueError:
                print("Please Enter An Integer")
                continue
            if ExcelChoice > len(Excel):
                print("Sorry A 'Valid' Interger Please")
                continue
            else:
                Excel = Excel[ExcelChoice]
                break

    return fileList(PDFs, Excel)

def parsePDF(currPDF):
    tableData = []
    with open (currPDF, "rb") as f:
      pdf = pdftotext.PDF(f)
    print "Now Reading ", currPDF
    for i in range(0,len(pdf)):
        tableData.append(extractTableData(pdf[i]))

    return tableData

def extractTableData(currPage):
    tableRow = []
    rawLines = currPage.split('\n')

    #Start and End Date
    ParsingString = rawLines[0].split()
    tableRow.extend((ParsingString[4],ParsingString[6]))

    #Group Rating Areas
    ParsingString = rawLines[2].split()
    ParsingString = ParsingString[2][4:]
    tableRow.append(ParsingString)

    #Plan Name and State
    ParsingString = rawLines[3].split()
    ParsingString = ParsingString[5:]
    ParsingString = " ".join(ParsingString)
    tableRow.insert(-1, ParsingString)
    tableRow.insert(-1, ParsingString[:2])

    #First Row
    ParsingString = rawLines[5].split()
    tableRow.extend((ParsingString[1], ParsingString[1]))
    FirstColumn = len(tableRow)
    tableRow.append(ParsingString[3])
    SecondColumn = len(tableRow)
    tableRow.append(ParsingString[5])

    #Loop Through Rows
    for i in range(6,20):
        ParsingString = rawLines[i].split()
        tableRow.insert(FirstColumn, ParsingString[1])
        tableRow.insert(SecondColumn, ParsingString[3])
        if i < 19:
            tableRow.append(ParsingString[5])
            FirstColumn += 1
            SecondColumn += 1
        else:
            tableRow.extend((ParsingString[5], ParsingString[5]))
    return "~".join(tableRow)

def writeToTemplate(tableData):
    print("Now Writing The Data To The Template")

    #Load Template File
    TemplateName = tableData.pop(0)
    WorkBook = load_workbook(TemplateName)
    Sheet = WorkBook.active

    #Iterate Through Data and Write To Cells
    for currRow, currRowData in enumerate(tableData, start=2):
        RowVals = currRowData.split('~')
        for col, item in enumerate(RowVals, start=1):
            Sheet.cell(row=currRow, column=col).value = item

    #Prompt The User To Save The Table Under A Different Name (Saves The Template File)
    NamePrompt = raw_input("Would You Like To Save The File With A Different Name? [Y/N]: ").upper()
    if NamePrompt.startswith('Y'):
        TemplateName = raw_input("What Is The New Name? (No Extension Needed): ") + ".xlsx"
        print("Okay Saving As: " + TemplateName)
    WorkBook.save(TemplateName)


#### Main ###
def main():
    listFiles = checkForFiles()
    tableData = [listFiles.Excel]
    for i in listFiles.PDFs:
        tableData = tableData + parsePDF(i)
    writeToTemplate(tableData)
    EndProg = raw_input("All Done!\nPress [Enter] to Quit ")

main()
